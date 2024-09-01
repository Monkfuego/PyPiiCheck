import PyPDF2
import openpyxl
import pytesseract
from PIL import Image
import re
import os
import cv2
from gov_id_detector.regex_patterns import PATTERNS, TEMPLATES
import numpy as np

# Define the documents that require both template and regex matches
STRICT_DOCUMENTS = ['Aadhar Card','Pan Card','Voter Id','Passport','DriverLicense','NREGA Job Card']

def extract_text_from_pdf(file_path):
    with open(file_path, 'rb') as file:
        reader = PyPDF2.PdfReader(file)
        text = ''.join([page.extract_text() for page in reader.pages])
    return text

def extract_text_from_image(file_path):
    image = Image.open(file_path)
    text = pytesseract.image_to_string(image)
    return text

def extract_text_from_excel(file_path):
    dataframe = openpyxl.load_workbook(file_path)
    dataframe1 = dataframe.active
    text = ''
    for row in range(0, dataframe1.max_row):
        for col in dataframe1.iter_cols(1, dataframe1.max_column):
            text += str(col[row].value)
    return text

def extract_text(file_path):
    if file_path.endswith('.pdf'):
        return extract_text_from_pdf(file_path)
    elif file_path.endswith('.docx'):
        # return extract_text_from_docx(file_path)
        pass
    elif file_path.endswith(('.png', '.jpg', '.jpeg')):
        return extract_text_from_image(file_path)
    elif file_path.endswith((".xls" , ".xlsx" , ".xlsm" , ".xlsb" , ".xltm" , ".xltx" , ".csv")):
        return extract_text_from_excel(file_path)

def match_template(image_path, template_path):
    image = cv2.imread(image_path)
    template = cv2.imread(template_path)
    result = cv2.matchTemplate(image, template, cv2.TM_CCOEFF_NORMED)
    min_val, max_val, min_loc, max_loc = cv2.minMaxLoc(result)
    if max_val > 0.4:
        return True
    return False

def check_templates(image_path):
    matches = {}
    for template_name, template_path in TEMPLATES.items():
        if match_template(image_path, template_path):
            matches[template_name] = True
    return matches

def check_patterns(text):
    matches = {}
    for pattern_name, pattern in PATTERNS.items():
        match = pattern.search(text)
        if match:
            matches[pattern_name] = match.group()
    return matches

def process_file(file_path):
    if file_path.endswith(('.png', '.jpg', '.jpeg')):
        template_matches = check_templates
